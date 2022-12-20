import csv
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


def get_salary_avg(dct):
    for key, value in dct.items():
        dct[key] = 0 if len(value) == 0 else int(sum(value) / len(value))
    return dct


class Vacancy:
    """
    Класс представления параметров вакансии

    Attributes:
        currency_to_rub (dict) : Словарь для перевода оклада в рубли
        name (str): Название вакансии
        salary_from (str / int / float): Нижняя граница вилки оклада
        salary_to (str / int / float): Верхняя граница вилки оклада
        salary_currency (str): Код валюты оклада
        area_name (str): Регион
        published_at (str): Дата публикации вакансии

    """
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                       "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

    def __init__(self, dct):
        """Инициализатор класса Vacancy
        :param self:
            dct (dict):
                dct['name'] (str): Название вакансии
                dct['salary_from'] (str / int / float): Нижняя граница вилки оклада
                dct['salary_to'] (str / int / float): Верхняя граница вилки оклада
                dct['salary_currency'] (str): Код валюты оклада
                dct['area_name'] (str): Регион
                dct['published_at'] (str): Дата публикации вакансии
        """
        self.name = dct['name']
        self.salary_from = float(dct['salary_from'])
        self.salary_to = float(dct['salary_to'])
        self.salary_currency = dct['salary_currency']
        self.area_name = dct['area_name']
        self.published_at = dct['published_at']

    def get_average_rub_salary(self):
        """Метод возвращает среднюю зарплату оклада в рублях

        :return:
            (float): Средняя зарплата оклада в рублях
        """
        return 0.5 * (self.salary_from + self.salary_to) * self.currency_to_rub[self.salary_currency]

    def get_published_vacancy_year(self):
        """Возвращает год публикации вакансии
        :return:
        (int): Год публикации вакансии
        """
        return int(self.published_at[:4])


class DataSet:
    """Класс для представления данных о вакансиях
    :Attributes:
        file_name (str): Название файла для анализа
        vacancy_name (str): Название выбранной вакансии
    """
    def __init__(self, file_name, name):
        """Инициализирует клас DataSet
        :param file_name:
            (str): Название файла
        :param name:
            (str): Название выбранной вакансии
        """
        self.file_name = file_name
        self.vacancy_name = name

    def parse_csv(self):
        """Парсит сырые данных из файла по словарям
        :return:
            salary (dict): Зарплата по годам
            amount (dict): Количество вакансий по годам
            vacancy_salary (dict): Зарплата по годам для выбранной вакансии
            vacancy_amount (dict): Количество вакансий по годам для выбранной вакансии
            salary_city (dict): Зарплата по городам
            share_city(dict): Количество по городам
            count (int): Общее количество всех вакансий
        """
        salary = dict()
        amount = dict()
        vacancy_salary = dict()
        vacancy_amount = dict()
        salary_city = dict()
        share_city = dict()
        count = 0
        with open(self.file_name, encoding='utf-8-sig') as vacancies:
            rows = list(csv.reader(vacancies))
            titles = rows[0]
            for row in rows[1:]:
                if not (len(row) == len(titles) and all(row)):
                    continue
                vacancy = Vacancy(dict(zip(titles, row)))
                count += 1
                mean_salary = vacancy.get_average_rub_salary()
                city = vacancy.area_name
                year = vacancy.get_published_vacancy_year()
                if year not in salary:
                    salary[year] = [mean_salary]
                else:
                    salary[year] += [mean_salary]

                if year not in amount:
                    amount[year] = 1
                else:
                    amount[year] += 1

                if city not in salary_city:
                    salary_city[city] = [mean_salary]
                else:
                    salary_city[city] += [mean_salary]

                if city not in share_city:
                    share_city[city] = 1
                else:
                    share_city[city] += 1

                if self.vacancy_name in vacancy.name:
                    if year not in vacancy_salary:
                        vacancy_salary[year] = [mean_salary]
                    else:
                        vacancy_salary[year] += [mean_salary]
                    if year not in vacancy_amount:
                        vacancy_amount[year] = 1
                    else:
                        vacancy_amount[year] += 1

            return salary, amount, vacancy_salary, vacancy_amount, salary_city, share_city, count

    def get_clear_data(self):
        """Преобразует сырые данные из метода parse_csv
        :return:
            salary (dict): Средняя зарплата по годам
            amount (dict): Количество вакансий по годам
            vacancy_salary (dict): Средняя зарплата по годам для выбранной вакансии
            vacancy_amount (dict): Количество вакансий по годам для выбранной вакансии
            salary_city (dict): Зарплата по городам (В порядке убывания)
            share_city(dict): Доля вакансий по городам (В порядке убывания)
        """
        salary, amount, vacancy_salary, vacancy_amount, salary_city, share_city, count = self.parse_csv()
        salary = get_salary_avg(salary)
        vacancy_salary = get_salary_avg(vacancy_salary)
        for k, v in share_city.items():
            share_city[k] = round(v / count, 4)
        share_city = list(filter(lambda x: x[-1] > 0.01, [(key, value) for key, value in share_city.items()]))

        salary_city = dict(
            sorted([(key, value) for key, value in get_salary_avg(salary_city).items() if key in dict(share_city)],
                   key=lambda x: x[-1], reverse=True)[:10])

        return salary, amount, vacancy_salary, vacancy_amount, salary_city, dict(
            sorted(share_city, key=lambda x: x[-1], reverse=True)[:10])


class Report:
    """Класс для представления отчётов
    Attributes:
        wb (Workbook): Workbook is the container for all other parts of the document
        vacancy_name (str): Название выбранной вакансии
        salary (dict): Средняя зарплата по годам
        amount (dict): Количество вакансий по годам
        this_vacancy_salary (dict): Средняя зарплата по годам для выбранной вакансии
        this_vacancy_amount (dict): Количество вакансий по годам для выбранной вакансии
        salary_city (dict): Зарплата по городам (В порядке убывания)
        share_city(dict): Доля вакансий по городам (В порядке убывания)
    """
    def __init__(self, vacancy_name, salary, amount, this_vacancy_salary, this_vacancy_amount, salary_city, share_city):
        """Инициализирует класс Report
        :param:
            vacancy_name (str): Название выбранной вакансии
        :param:
            salary (dict): Средняя зарплата по годам
        :param:
            amount (dict): Количество вакансий по годам
        :param:
            this_vacancy_salary (dict): Средняя зарплата по годам для выбранной вакансии
        :param:
            this_vacancy_amount (dict): Количество вакансий по годам для выбранной вакансии
        :param:
            salary_city (dict): Зарплата по городам (В порядке убывания)
        :param:
            share_city(dict): Доля вакансий по городам (В порядке убывания)
        """
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.salary = salary
        self.amount = amount
        self.this_vacancy_salary = this_vacancy_salary
        self.this_vacancy_amount = this_vacancy_amount
        self.salary_city = salary_city
        self.share_city = share_city

    def generate_image(self):
        """Формирует отчет в виде graph.png с графиками"""
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        bar1 = ax1.bar(np.array(list(self.salary.keys())) - 0.4, self.salary.values(), width=0.4)
        bar2 = ax1.bar(np.array(list(self.salary.keys())), self.this_vacancy_salary.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list(self.salary.keys())) - 0.2, list(self.salary.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(np.array(list(self.amount.keys())) - 0.4, self.amount.values(), width=0.4)
        bar2 = ax2.bar(np.array(list(self.amount.keys())), self.this_vacancy_amount.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                   prop={'size': 8})
        ax2.set_xticks(np.array(list(self.amount.keys())) - 0.2, list(self.amount.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.salary_city.keys()))]),
                 list(reversed(list(self.salary_city.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.share_city.values()])
        ax4.pie(list(self.share_city.values()) + [other], labels=list(self.share_city.keys()) + ['Другие'],
                textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self):
        """Формирует отчет в виде pdf-файла со статистикой и графиками"""
        template = Environment(loader=FileSystemLoader('templates')).get_template("pdf.html")
        statistic = []
        for year in self.salary:
            statistic.append([year, self.salary[year], self.this_vacancy_salary[year], self.amount[year],
                              self.this_vacancy_amount[year]])
        for key, value in self.share_city.items():
            self.share_city[key] = str(round(value * 100, 2)) + '%'
        pdf = template.render({'name': dataset.vacancy_name,
                               'path': r'C:\Users\ilyam\PycharmProjects\pythonProject\graph.png',
                               'statistic': statistic, 'salary_city': self.salary_city,
                               'share_city': self.share_city})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def generate_excel(self):
        """Формирует статистику в виде .xlsx"""
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
                    'Количество вакансий - ' + self.vacancy_name])
        for year in self.salary:
            ws1.append([year, self.salary[year], self.this_vacancy_salary[year], self.amount[year],
                        self.this_vacancy_amount[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.salary_city.items(), self.share_city.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = font_bold
            ws2[col + '1'].font = font_bold

        for index, _ in enumerate(self.salary_city):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row, _ in enumerate(self.salary):
            for col in 'ABCDE':
                ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)
        self.wb.save('report.xlsx')


vacancy_name = 'Введите название IT-профессии'
dataset = DataSet('vacancies_by_year.csv', vacancy_name)
report = Report(vacancy_name, *dataset.get_clear_data())
choice = input('Отчет, вакансии или статистика?')
if choice == 'Отчет':
    report.generate_pdf()
elif choice == 'Вакансии':
    report.generate_image()
elif choice == 'Статистика':
    report.generate_excel()
else:
    print('Неправильный формат ввода')
